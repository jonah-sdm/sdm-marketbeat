"""
Excel Export helpers
────────────────────
Creates formatted .xlsx files from structured sheet data.
Used by all three sectors for download/export functionality.
"""

import os
import tempfile


# Create a formatted Excel (.xlsx) file from a list of sheet definitions.
# Each sheet is a dict with "title", "headers" (list of column names), and
# "rows" (list of lists). Returns the path to the saved temp file.
def create_xlsx_from_data(sheets_data, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1B2A4A"
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    normal = Font(name="Calibri", size=10)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = openpyxl.Workbook()

    for si, sd in enumerate(sheets_data):
        ws = wb.active if si == 0 else wb.create_sheet()
        ws.title = sd["title"][:31]  # Excel 31 char limit

        # Headers
        for col, header in enumerate(sd["headers"], 1):
            c = ws.cell(row=1, column=col, value=header)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = thin_border

        # Data rows
        for ri, row_data in enumerate(sd["rows"]):
            for col, value in enumerate(row_data, 1):
                c = ws.cell(row=ri + 2, column=col, value=value)
                c.font = normal
                c.border = thin_border
                c.alignment = wrap
            if ri % 2 == 1:
                for col in range(1, len(sd["headers"]) + 1):
                    ws.cell(row=ri + 2, column=col).fill = alt_fill

        # Auto-width (approximate)
        for col in range(1, len(sd["headers"]) + 1):
            max_len = len(str(sd["headers"][col - 1]))
            for row_data in sd["rows"][:50]:
                if col - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col - 1] or "")))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

    path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(path)
    return path
