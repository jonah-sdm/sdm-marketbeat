#!/usr/bin/env python3
"""Convert legal_search_results.json to a formatted Excel spreadsheet.

Sheets:
  1. Summary       — source counts, entity counts, manual URLs
  2. Case Details  — full case data per source
  3. Parties       — all unique parties classified by type
  4. Companies     — unique companies across all sources
  5. Law Firms     — unique law firms
  6. Attorneys     — unique attorneys
  7. Judges        — unique judges
"""

import json
import sys
import subprocess

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "-q", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────

NAVY = "1B2A4A"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
GREEN = "27AE60"
RED = "E74C3C"
ORANGE = "F39C12"
PURPLE = "8E44AD"

title_font   = Font(name="Calibri", size=14, bold=True, color=NAVY)
header_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill  = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
alt_fill     = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
section_font = Font(name="Calibri", size=12, bold=True, color=NAVY)
section_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
link_font    = Font(name="Calibri", size=10, color="0563C1", underline="single")
normal       = Font(name="Calibri", size=10)
bold         = Font(name="Calibri", size=10, bold=True)
ok_font      = Font(name="Calibri", size=10, bold=True, color=GREEN)
err_font     = Font(name="Calibri", size=10, bold=True, color=RED)
warn_font    = Font(name="Calibri", size=10, bold=True, color=ORANGE)
thin_border  = Border(bottom=Side(style="thin", color="CCCCCC"))
center       = Alignment(horizontal="center", vertical="center")
wrap         = Alignment(vertical="top", wrap_text=True)


def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers):
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, thin_border


def styled_cell(ws, row, col, value, font=normal, align=None, hyper=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.border = thin_border
    if align:
        c.alignment = align
    if hyper:
        c.hyperlink = hyper
        c.font = link_font
    return c


def write_entity_sheet(wb, title, tab_color, headers, widths, items, row_fn):
    """Generic helper to create an entity listing sheet."""
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    col_widths(ws, widths)
    ws.merge_cells(f"B2:{get_column_letter(len(headers))}2")
    ws.cell(row=2, column=2, value=f"{title} ({len(items)})").font = title_font
    header_row(ws, 4, headers)
    for i, item in enumerate(items, 1):
        r = 4 + i
        row_fn(ws, r, i, item)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = alt_fill
    ws.freeze_panes = "A5"
    return ws


# ── Load data ─────────────────────────────────────────────────────────────

input_file = sys.argv[1] if len(sys.argv) > 1 else "legal_search_results.json"
with open(input_file) as f:
    data = json.load(f)

entities = data.get("entities", {})
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════
# Sheet 1: Summary
# ══════════════════════════════════════════════════════════════════════════

ws = wb.active
ws.title = "Summary"
ws.sheet_properties.tabColor = NAVY
col_widths(ws, [4, 32, 14, 16, 50, 55])

ws.merge_cells("B2:F2")
ws.cell(row=2, column=2, value=f"Legal Case Search: \"{data['query']}\"").font = title_font
search_date = data.get("search_date", "")[:10] or "N/A"
ws.cell(row=3, column=2, value=f"Search date: {search_date}").font = normal
ws.cell(row=3, column=4, value=f"Total from APIs: {data.get('total_from_apis', 0)}").font = bold

# Entity counts row
ec = data.get("entity_counts", {})
row = 4
if ec:
    counts_str = (f"Parties: {ec.get('parties',0)}  |  Companies: {ec.get('companies',0)}  |  "
                  f"Law Firms: {ec.get('law_firms',0)}  |  Attorneys: {ec.get('attorneys',0)}  |  "
                  f"Judges: {ec.get('judges',0)}")
    ws.cell(row=row, column=2, value=counts_str).font = bold

# API Results section
row = 6
ws.merge_cells(f"B{row}:F{row}")
ws.cell(row=row, column=2, value="API Sources").font = section_font
ws.cell(row=row, column=2).fill = section_fill
row += 1
header_row(ws, row, ["", "Source", "Cases Found", "Status", "Notes", "Search URL"])

for i, src in enumerate(data.get("api_results", [])):
    row += 1
    styled_cell(ws, row, 2, src["source"])
    count = src.get("count")
    styled_cell(ws, row, 3, count if count is not None else "N/A",
                font=bold if count and count > 0 else normal, align=center)
    if src.get("error"):
        styled_cell(ws, row, 4, "Error", font=err_font, align=center)
        styled_cell(ws, row, 5, src["error"])
    elif count == 0:
        styled_cell(ws, row, 4, "No results", font=warn_font, align=center)
        styled_cell(ws, row, 5, "")
    else:
        styled_cell(ws, row, 4, "OK", font=ok_font, align=center)
        styled_cell(ws, row, 5, f"{count} result(s) via API")
    if src.get("url"):
        styled_cell(ws, row, 6, src["url"], hyper=src["url"])
    if i % 2 == 1:
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = alt_fill

# Manual URLs section
row += 2
ws.merge_cells(f"B{row}:F{row}")
ws.cell(row=row, column=2, value="Manual Search Sources").font = section_font
ws.cell(row=row, column=2).fill = section_fill
row += 1
header_row(ws, row, ["", "Source", "Type", "Status", "Notes", "Search URL"])
for i, src in enumerate(data.get("manual_search_urls", [])):
    row += 1
    styled_cell(ws, row, 2, src["source"])
    styled_cell(ws, row, 3, "Manual", align=center)
    styled_cell(ws, row, 4, "URL ready", font=warn_font, align=center)
    styled_cell(ws, row, 5, src.get("note", ""))
    styled_cell(ws, row, 6, src["url"], hyper=src["url"])
    if i % 2 == 1:
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = alt_fill
ws.freeze_panes = "A8"


# ══════════════════════════════════════════════════════════════════════════
# Sheet 2: Case Details
# ══════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Case Details")
ws2.sheet_properties.tabColor = GREEN
col_widths(ws2, [4, 6, 42, 20, 20, 30, 14, 14, 14, 20, 20, 30, 30, 55])

ws2.merge_cells("B2:N2")
ws2.cell(row=2, column=2, value=f"Detailed Case Data: \"{data['query']}\"").font = title_font

row = 4
case_num = 0
for src in data.get("api_results", []):
    if not src.get("cases"):
        continue

    ws2.merge_cells(f"B{row}:N{row}")
    ws2.cell(row=row, column=2, value=f"{src['source']}  ({src.get('count', '?')} total)").font = section_font
    ws2.cell(row=row, column=2).fill = section_fill
    row += 1

    if "SEC" in src["source"]:
        headers = ["", "#", "Company", "Entity (raw)", "Form Type", "Filing Description",
                   "File Date", "Period", "", "", "", "", "", "Source URL"]
        header_row(ws2, row, headers)
        row += 1
        for i, c in enumerate(src["cases"], 1):
            case_num += 1
            styled_cell(ws2, row, 2, case_num, align=center)
            styled_cell(ws2, row, 3, c.get("company", ""), align=wrap)
            styled_cell(ws2, row, 4, c.get("entity", ""))
            styled_cell(ws2, row, 5, c.get("form_type", ""), align=center)
            styled_cell(ws2, row, 6, c.get("file_description", ""), align=wrap)
            styled_cell(ws2, row, 7, c.get("file_date", ""), align=center)
            styled_cell(ws2, row, 8, c.get("period_of_report", ""), align=center)
            if src.get("url"):
                styled_cell(ws2, row, 14, src["url"], hyper=src["url"])
            if i % 2 == 0:
                for col in range(1, 15):
                    ws2.cell(row=row, column=col).fill = alt_fill
            row += 1
    else:
        headers = ["", "#", "Case Name", "Plaintiff", "Defendant", "Court",
                   "Filed", "Terminated", "Docket #", "Case Type / Cause",
                   "Judge", "Law Firms", "Attorneys", "URL"]
        header_row(ws2, row, headers)
        row += 1
        for i, c in enumerate(src["cases"], 1):
            case_num += 1
            styled_cell(ws2, row, 2, case_num, align=center)
            styled_cell(ws2, row, 3, c.get("name", ""), align=wrap)
            styled_cell(ws2, row, 4, c.get("plaintiff", ""), align=wrap)
            styled_cell(ws2, row, 5, c.get("defendant", ""), align=wrap)
            styled_cell(ws2, row, 6, c.get("court", ""), align=wrap)
            styled_cell(ws2, row, 7, c.get("date_filed", ""), align=center)
            styled_cell(ws2, row, 8, c.get("date_terminated", ""), align=center)
            styled_cell(ws2, row, 9, c.get("docket_number", ""), align=center)
            case_type = c.get("case_type", "") or c.get("cause", "")
            styled_cell(ws2, row, 10, case_type, align=wrap)
            styled_cell(ws2, row, 11, c.get("judge", ""))
            firms = c.get("law_firms", [])
            styled_cell(ws2, row, 12, ", ".join(firms) if firms else "", align=wrap)
            attys = c.get("attorneys", [])
            styled_cell(ws2, row, 13, ", ".join(attys[:5]) if attys else "", align=wrap)
            docket_url = c.get("docket_url", c.get("opinion_url", src.get("url", "")))
            if docket_url:
                styled_cell(ws2, row, 14, docket_url, hyper=docket_url)
            if i % 2 == 0:
                for col in range(1, 15):
                    ws2.cell(row=row, column=col).fill = alt_fill
            row += 1
    row += 1

ws2.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════
# Sheets 3-7: Entity sheets
# ══════════════════════════════════════════════════════════════════════════

# Parties
parties = entities.get("parties", [])
if parties:
    def party_row(ws, r, i, item):
        styled_cell(ws, r, 2, i, align=center)
        styled_cell(ws, r, 3, item.get("name", ""))
        styled_cell(ws, r, 4, item.get("type", ""), align=center)
        styled_cell(ws, r, 5, ", ".join(item.get("sources", [])), align=wrap)
    write_entity_sheet(wb, "Parties", ORANGE,
                       ["", "#", "Party Name", "Type", "Found In"],
                       [4, 6, 50, 16, 40], parties, party_row)

# Companies
companies = entities.get("companies", [])
if companies:
    def co_row(ws, r, i, item):
        styled_cell(ws, r, 2, i, align=center)
        styled_cell(ws, r, 3, item.get("name", ""))
        styled_cell(ws, r, 4, ", ".join(item.get("sources", [])), align=wrap)
    write_entity_sheet(wb, "Companies", GREEN,
                       ["", "#", "Company Name", "Found In"],
                       [4, 6, 55, 45], companies, co_row)

# Law Firms
firms = entities.get("law_firms", [])
if firms:
    def firm_row(ws, r, i, item):
        styled_cell(ws, r, 2, i, align=center)
        styled_cell(ws, r, 3, item.get("name", ""))
        styled_cell(ws, r, 4, ", ".join(item.get("sources", [])), align=wrap)
    write_entity_sheet(wb, "Law Firms", PURPLE,
                       ["", "#", "Firm Name", "Found In"],
                       [4, 6, 60, 40], firms, firm_row)

# Attorneys
attorneys = entities.get("attorneys", [])
if attorneys:
    def atty_row(ws, r, i, item):
        styled_cell(ws, r, 2, i, align=center)
        styled_cell(ws, r, 3, item.get("name", ""))
        styled_cell(ws, r, 4, ", ".join(item.get("sources", [])), align=wrap)
    write_entity_sheet(wb, "Attorneys", "2980B9",
                       ["", "#", "Attorney Name", "Found In"],
                       [4, 6, 45, 45], attorneys, atty_row)

# Judges
judges = entities.get("judges", [])
if judges:
    def judge_row(ws, r, i, item):
        styled_cell(ws, r, 2, i, align=center)
        styled_cell(ws, r, 3, item.get("name", ""))
        styled_cell(ws, r, 4, ", ".join(item.get("sources", [])), align=wrap)
    write_entity_sheet(wb, "Judges", "C0392B",
                       ["", "#", "Judge Name", "Found In"],
                       [4, 6, 40, 45], judges, judge_row)


# ── Save ──────────────────────────────────────────────────────────────────

output = sys.argv[2] if len(sys.argv) > 2 else input_file.replace(".json", ".xlsx")
wb.save(output)
print(f"Saved: {output}")
print(f"  Sheets: {', '.join(wb.sheetnames)}")
print(f"  Cases: {case_num} | Parties: {len(parties)} | Companies: {len(companies)} "
      f"| Firms: {len(firms)} | Attorneys: {len(attorneys)} | Judges: {len(judges)}")
