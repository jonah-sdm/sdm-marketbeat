#!/usr/bin/env python3
"""
SDM Industry Search Engine
───────────────────────────
Unified Flask application with three intelligence sectors:
  1. Legal & Litigation Intelligence
  2. Digital Asset Treasury & Finance
  3. Event Intelligence

Usage:
    python3 app.py [port]          # default port 5055
"""

import csv
import io
import json
import os
import re
import subprocess
import sys
import tempfile
import threading
import time
import urllib.parse

# ── Auto-install dependencies (local dev only; skipped on Render) ────────
if os.environ.get("RENDER") is None:
    _REQUIRED = [
        "flask", "requests", "beautifulsoup4", "openpyxl",
        "python-dotenv", "requests-html",
    ]
    for pkg in _REQUIRED:
        _import = pkg.replace("beautifulsoup4", "bs4").replace("python-dotenv", "dotenv").replace("requests-html", "requests_html")
        try:
            __import__(_import)
        except ImportError:
            print(f"Installing {pkg}...")
            subprocess.check_call([sys.executable, "-m", "pip", "-q", "install", pkg])

from datetime import datetime
from dotenv import load_dotenv
from flask import (
    Flask, Blueprint, request, render_template, send_file, jsonify, redirect,
)
import requests as http_requests
from bs4 import BeautifulSoup

# ── Legal search backend ────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from legal_case_search import (
    make_session, search_courtlistener_opinions, search_courtlistener_dockets,
    search_sec_edgar, search_canlii, generate_search_urls, collect_plaintiffs,
    aggregate_entities,
)

# ── Config ───────────────────────────────────────────────────────────────
load_dotenv()
APOLLO_API_KEY = os.getenv("APOLLO_API_KEY", "").strip()
API_BASE = "https://api.apollo.io/api/v1"

TREASURY_TITLES = [
    "Treasurer", "Treasury Manager", "Treasury Director", "Head of Treasury",
    "VP Treasury", "Treasury Analyst", "Treasury Operations",
    "Chief Financial Officer", "CFO", "Finance Director", "Head of Finance",
    "VP Finance", "Controller", "Director of Finance", "Financial Controller",
]

TITLE_SETS = {
    "treasury": TREASURY_TITLES[:7],
    "cfo": TREASURY_TITLES[7:],
    "all": TREASURY_TITLES,
}


# ═════════════════════════════════════════════════════════════════════════
# Apollo API helpers
# ═════════════════════════════════════════════════════════════════════════

def _check_api_key():
    if not APOLLO_API_KEY or APOLLO_API_KEY == "your_key_here":
        return False, (
            "Apollo.io API key not configured. "
            "Edit your .env file and set APOLLO_API_KEY=your_actual_key"
        )
    return True, ""


def _apollo_headers():
    return {"Content-Type": "application/json", "X-Api-Key": APOLLO_API_KEY}


def _apollo_post(endpoint, body):
    url = f"{API_BASE}/{endpoint}"
    try:
        r = http_requests.post(url, json=body, headers=_apollo_headers(), timeout=30)
        if r.status_code == 401:
            return None, "Invalid API key"
        if r.status_code == 403:
            return None, f"Endpoint not accessible: {endpoint}"
        if r.status_code == 429:
            return None, "Rate limit exceeded — wait and retry"
        r.raise_for_status()
        return r.json(), None
    except http_requests.exceptions.HTTPError as e:
        return None, f"HTTP {e.response.status_code}"
    except http_requests.exceptions.ConnectionError:
        return None, "Cannot reach Apollo.io"
    except http_requests.exceptions.Timeout:
        return None, "Request timed out"
    except Exception as e:
        return None, str(e)


def match_person_at_company(company_name, company_domain, title):
    body = {
        "organization_name": company_name,
        "title": title,
        "reveal_personal_emails": False,
    }
    if company_domain:
        body["organization_domain"] = company_domain

    data, err = _apollo_post("people/match", body)
    if err:
        return None
    person = data.get("person")
    if not person:
        return None

    org = person.get("organization", {}) or {}
    name = person.get("name", "")
    first = person.get("first_name", "")
    if not name and not first:
        name = "(name redacted — paid plan)"

    emp = person.get("employment_history", [])
    current_role = ""
    if emp:
        for e in emp:
            if e.get("current"):
                current_role = e.get("title", "")
                break

    return {
        "name": name,
        "first_name": first,
        "last_name": person.get("last_name", ""),
        "title": person.get("title", current_role or title),
        "headline": person.get("headline", ""),
        "seniority": person.get("seniority", ""),
        "departments": person.get("departments", []),
        "functions": person.get("functions", []),
        "email": person.get("email", ""),
        "email_status": person.get("email_status", ""),
        "phone": person.get("phone_number", ""),
        "linkedin_url": person.get("linkedin_url", ""),
        "city": person.get("city", ""),
        "state": person.get("state", ""),
        "country": person.get("country", ""),
        "company": org.get("name", company_name),
        "company_domain": org.get("primary_domain", company_domain),
        "company_website": org.get("website_url", ""),
        "company_industry": org.get("industry", ""),
        "company_size": org.get("estimated_num_employees", ""),
        "company_founded": org.get("founded_year", ""),
        "apollo_id": person.get("id", ""),
        "confirmed_exists": True,
    }


def find_people_at_companies(companies, titles, max_companies=10):
    people = []
    seen = set()
    for co in companies[:max_companies]:
        for title in titles:
            person = match_person_at_company(co["name"], co.get("domain", ""), title)
            if not person:
                continue
            uid = person.get("apollo_id") or f"{person.get('name','')}|{person.get('title','')}"
            if uid in seen:
                continue
            seen.add(uid)
            people.append(person)
            time.sleep(0.15)
    return people


def enrich_person_by_name(name, company=None):
    """Try to find a person by name (and optionally company) via Apollo."""
    body = {"reveal_personal_emails": False}
    parts = name.strip().split()
    if len(parts) >= 2:
        body["first_name"] = parts[0]
        body["last_name"] = " ".join(parts[1:])
    else:
        body["first_name"] = name
        body["last_name"] = ""
    if company:
        body["organization_name"] = company

    data, err = _apollo_post("people/match", body)
    if err or not data:
        return None
    person = data.get("person")
    if not person:
        return None

    org = person.get("organization", {}) or {}
    return {
        "name": person.get("name", name),
        "first_name": person.get("first_name", ""),
        "last_name": person.get("last_name", ""),
        "title": person.get("title", ""),
        "headline": person.get("headline", ""),
        "email": person.get("email", ""),
        "email_status": person.get("email_status", ""),
        "phone": person.get("phone_number", ""),
        "linkedin_url": person.get("linkedin_url", ""),
        "city": person.get("city", ""),
        "state": person.get("state", ""),
        "country": person.get("country", ""),
        "company": org.get("name", company or ""),
        "company_domain": org.get("primary_domain", ""),
        "company_website": org.get("website_url", ""),
        "company_industry": org.get("industry", ""),
        "apollo_id": person.get("id", ""),
    }


# ═════════════════════════════════════════════════════════════════════════
# Bitcoin Treasuries API
# ═════════════════════════════════════════════════════════════════════════

BT_API = "https://playground.bitcointreasuries.net/api"
_bt_cache = {"companies": None, "btc_price": None, "ts": 0}
BT_CACHE_TTL = 86400  # 24 hours


def _bt_get(path, retries=2):
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) SDM-Search/1.0",
        "Accept": "application/json",
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            r = http_requests.get(f"{BT_API}/{path}", headers=headers, timeout=15)
            try:
                data = r.json()
                if data:
                    return data, None
            except ValueError:
                pass
            if r.status_code >= 400 and attempt < retries:
                time.sleep(1)
                continue
            if r.status_code >= 400:
                return None, f"HTTP {r.status_code}"
            return r.json(), None
        except Exception as e:
            last_err = e
            if attempt < retries:
                time.sleep(1)
    return None, str(last_err)


def fetch_bt_companies():
    now = time.time()
    if _bt_cache["companies"] and (now - _bt_cache["ts"]) < BT_CACHE_TTL:
        return _bt_cache["companies"], _bt_cache["btc_price"], None

    data, err = _bt_get("companies")
    if err:
        return None, None, f"Bitcoin Treasuries API error: {err}"

    btc_price = None
    price_data, _ = _bt_get("market/btc-price")
    if price_data:
        btc_price = price_data.get("price")

    _bt_cache["companies"] = data
    _bt_cache["btc_price"] = btc_price
    _bt_cache["ts"] = now
    return data, btc_price, None


def format_btc_holdings(companies, btc_price):
    result = []
    for co in companies:
        btc = 0
        try:
            btc = int(co.get("btcHoldings", 0) or 0)
        except (ValueError, TypeError):
            try:
                btc = int(float(co.get("btcHoldings", 0) or 0))
            except (ValueError, TypeError):
                pass
        if btc <= 0:
            continue
        usd_val = btc * btc_price if btc_price else None
        mcap = None
        try:
            mcap = float(co.get("marketCap", 0) or 0)
        except (ValueError, TypeError):
            pass
        result.append({
            "name": co.get("name", ""),
            "symbol": co.get("symbol", ""),
            "country": co.get("country", ""),
            "btc_holdings": btc,
            "usd_value": usd_val,
            "market_cap": mcap,
            "stock_price": co.get("stockPrice"),
            "mnav": co.get("mNav"),
            "btc_yield": co.get("btcYield"),
            "avg_cost_basis": co.get("avgCostBasis"),
            "last_updated": co.get("lastUpdated", ""),
        })
    result.sort(key=lambda x: x["btc_holdings"], reverse=True)
    for i, co in enumerate(result):
        co["rank"] = i + 1
    return result


# ═════════════════════════════════════════════════════════════════════════
# Luma Event Scraping
# ═════════════════════════════════════════════════════════════════════════

_event_cache = {}
EVENT_CACHE_TTL = 1800  # 30 min


def scrape_luma_event(url):
    """Scrape event details and publicly visible attendees from a Luma event page."""
    cache_key = url
    if cache_key in _event_cache and (time.time() - _event_cache[cache_key]["ts"]) < EVENT_CACHE_TTL:
        return _event_cache[cache_key]["data"], None

    event_info = {
        "name": "",
        "date": "",
        "time": "",
        "location": "",
        "host": "",
        "description": "",
        "url": url,
    }
    attendees = []

    # Extract event slug for API approach
    slug = None
    m = re.search(r'lu\.ma/([^/?#]+)', url)
    if m:
        slug = m.group(1)

    # Try the Luma API first (faster, more reliable)
    if slug:
        api_event, api_attendees = _try_luma_api(slug)
        if api_event:
            event_info.update(api_event)
        if api_attendees:
            attendees = api_attendees

    # Fall back to HTML scraping if API didn't work
    if not attendees:
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
                "Accept": "text/html,application/xhtml+xml",
            }
            r = http_requests.get(url, headers=headers, timeout=20)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")

            # Try to extract event data from JSON-LD or meta tags
            for script in soup.find_all("script", type="application/ld+json"):
                try:
                    ld = json.loads(script.string)
                    if isinstance(ld, dict) and ld.get("@type") == "Event":
                        event_info["name"] = ld.get("name", event_info["name"])
                        event_info["description"] = ld.get("description", "")
                        loc = ld.get("location", {})
                        if isinstance(loc, dict):
                            event_info["location"] = loc.get("name", loc.get("address", ""))
                        start = ld.get("startDate", "")
                        if start:
                            event_info["date"] = start[:10]
                            event_info["time"] = start[11:16] if len(start) > 11 else ""
                        org = ld.get("organizer", {})
                        if isinstance(org, dict):
                            event_info["host"] = org.get("name", "")
                except (json.JSONDecodeError, TypeError):
                    pass

            # Extract from meta tags
            if not event_info["name"]:
                og_title = soup.find("meta", property="og:title")
                if og_title:
                    event_info["name"] = og_title.get("content", "")
            if not event_info["description"]:
                og_desc = soup.find("meta", property="og:description")
                if og_desc:
                    event_info["description"] = og_desc.get("content", "")

            # Try to find __NEXT_DATA__ for attendee info
            for script in soup.find_all("script", id="__NEXT_DATA__"):
                try:
                    next_data = json.loads(script.string)
                    props = next_data.get("props", {}).get("pageProps", {})

                    # Event data from Next.js
                    ev = props.get("event", props.get("initialData", {}).get("event", {}))
                    if ev:
                        event_info["name"] = ev.get("name", event_info["name"])
                        event_info["description"] = ev.get("description", event_info["description"])
                        geo = ev.get("geo_address_info", {})
                        if geo:
                            event_info["location"] = geo.get("full_address", geo.get("city", ""))
                        event_info["date"] = (ev.get("start_at", "") or "")[:10]

                    # Guest/attendee data
                    guests = props.get("guests", props.get("initialData", {}).get("guests", []))
                    if isinstance(guests, list):
                        for g in guests:
                            if isinstance(g, dict):
                                name = g.get("name", g.get("user_name", ""))
                                if not name:
                                    name = f"{g.get('first_name', '')} {g.get('last_name', '')}".strip()
                                if name:
                                    attendees.append({
                                        "name": name,
                                        "company": g.get("company", g.get("organization", "")),
                                        "role": g.get("job_title", g.get("title", "")),
                                        "avatar": g.get("avatar_url", ""),
                                    })
                except (json.JSONDecodeError, TypeError, KeyError):
                    pass

            # Try to find attendee names from visible HTML elements
            if not attendees:
                # Look for common attendee patterns in the page
                for el in soup.select('[class*="attendee"], [class*="guest"], [class*="participant"]'):
                    name_el = el.select_one('[class*="name"]')
                    if name_el and name_el.get_text(strip=True):
                        attendees.append({
                            "name": name_el.get_text(strip=True),
                            "company": "",
                            "role": "",
                        })

        except http_requests.exceptions.RequestException as e:
            if not event_info["name"]:
                return None, f"Failed to fetch event page: {str(e)}"

    # Try JS rendering as last resort if we have no attendees
    if not attendees:
        try:
            from requests_html import HTMLSession
            session = HTMLSession()
            r = session.get(url, timeout=20)
            r.html.render(timeout=15, sleep=2)

            # Re-parse with rendered HTML
            soup = BeautifulSoup(r.html.html, "html.parser")

            if not event_info["name"]:
                title_el = soup.select_one('h1, [class*="title"]')
                if title_el:
                    event_info["name"] = title_el.get_text(strip=True)

            for el in soup.select('[class*="attendee"], [class*="guest"], [class*="avatar"]'):
                name_el = el.select_one('[class*="name"]') or el
                name = name_el.get_text(strip=True)
                if name and len(name) > 1 and len(name) < 60:
                    attendees.append({
                        "name": name,
                        "company": "",
                        "role": "",
                    })
        except Exception:
            pass

    # Deduplicate attendees
    seen = set()
    unique = []
    for a in attendees:
        key = a["name"].lower().strip()
        if key and key not in seen:
            seen.add(key)
            unique.append(a)
    attendees = unique

    result = {"event": event_info, "attendees": attendees}
    _event_cache[cache_key] = {"data": result, "ts": time.time()}
    return result, None


def _try_luma_api(slug):
    """Try to get event data from Luma's public API."""
    event_info = {}
    attendees = []

    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "application/json",
        }

        # Try event API
        r = http_requests.get(f"https://api.lu.ma/event/get?event_slug={slug}",
                              headers=headers, timeout=15)
        if r.status_code == 200:
            data = r.json()
            ev = data.get("event", data)
            event_info = {
                "name": ev.get("name", ""),
                "description": ev.get("description", ""),
                "date": (ev.get("start_at", "") or "")[:10],
                "time": (ev.get("start_at", "") or "")[11:16],
                "location": "",
                "host": "",
            }
            geo = ev.get("geo_address_info", {})
            if geo:
                event_info["location"] = geo.get("full_address", geo.get("city", ""))
            hosts = data.get("hosts", [])
            if hosts:
                event_info["host"] = hosts[0].get("name", "")

        # Try guest list API
        r2 = http_requests.get(
            f"https://api.lu.ma/event/get-guests?event_slug={slug}&limit=200",
            headers=headers, timeout=15,
        )
        if r2.status_code == 200:
            guest_data = r2.json()
            entries = guest_data.get("entries", guest_data.get("guests", []))
            if isinstance(entries, list):
                for entry in entries:
                    guest = entry.get("guest", entry) if isinstance(entry, dict) else {}
                    user = entry.get("user", {}) if isinstance(entry, dict) else {}
                    name = (guest.get("name") or user.get("name") or
                            f"{user.get('first_name', '')} {user.get('last_name', '')}".strip())
                    if name:
                        attendees.append({
                            "name": name,
                            "company": guest.get("company", user.get("company", "")),
                            "role": guest.get("job_title", user.get("job_title", "")),
                            "avatar": user.get("avatar_url", ""),
                        })
    except Exception:
        pass

    return event_info, attendees


def enrich_attendees_with_apollo(attendees, max_enrich=50):
    """Enrich attendee list with Apollo.io contact data."""
    ok, _ = _check_api_key()
    if not ok:
        return attendees, 0

    enriched_count = 0
    for a in attendees[:max_enrich]:
        person = enrich_person_by_name(a["name"], a.get("company"))
        if person:
            a.update({
                "title": person.get("title", a.get("role", "")),
                "email": person.get("email", ""),
                "email_status": person.get("email_status", ""),
                "phone": person.get("phone", ""),
                "linkedin_url": person.get("linkedin_url", ""),
                "city": person.get("city", ""),
                "state": person.get("state", ""),
                "country": person.get("country", ""),
                "company": person.get("company", a.get("company", "")),
                "company_domain": person.get("company_domain", ""),
                "company_industry": person.get("company_industry", ""),
                "apollo_id": person.get("apollo_id", ""),
            })
            enriched_count += 1
            time.sleep(0.2)  # rate limit

    return attendees, enriched_count


# ═════════════════════════════════════════════════════════════════════════
# Excel Export Helpers
# ═════════════════════════════════════════════════════════════════════════

def _create_xlsx_from_data(sheets_data, filename):
    """Create an Excel file from a list of sheet definitions.
    Each sheet: {"title": str, "headers": [str], "rows": [[value, ...]]}
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1B2A4A"
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    normal = Font(name="Calibri", size=10)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = openpyxl.Workbook()

    for si, sd in enumerate(sheets_data):
        ws = wb.active if si == 0 else wb.create_sheet()
        ws.title = sd["title"][:31]  # Excel 31 char limit

        # Headers
        for col, header in enumerate(sd["headers"], 1):
            c = ws.cell(row=1, column=col, value=header)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = thin_border

        # Data rows
        for ri, row_data in enumerate(sd["rows"]):
            for col, value in enumerate(row_data, 1):
                c = ws.cell(row=ri + 2, column=col, value=value)
                c.font = normal
                c.border = thin_border
                c.alignment = wrap
            if ri % 2 == 1:
                for col in range(1, len(sd["headers"]) + 1):
                    ws.cell(row=ri + 2, column=col).fill = alt_fill

        # Auto-width (approximate)
        for col in range(1, len(sd["headers"]) + 1):
            max_len = len(str(sd["headers"][col - 1]))
            for row_data in sd["rows"][:50]:
                if col - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col - 1] or "")))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

    path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(path)
    return path


# ═════════════════════════════════════════════════════════════════════════
# Legal search backend
# ═════════════════════════════════════════════════════════════════════════

_legal_cache = {}
LEGAL_CACHE_TTL = 3600


def run_legal_search(company, sources=None):
    cache_key = f"{company}|{','.join(sorted(sources)) if sources else 'all'}"
    if cache_key in _legal_cache and (time.time() - _legal_cache[cache_key]["ts"]) < LEGAL_CACHE_TTL:
        return _legal_cache[cache_key]["data"]
    source_map = {
        "opinions": search_courtlistener_opinions,
        "dockets": search_courtlistener_dockets,
        "sec": search_sec_edgar,
        "canlii": search_canlii,
    }
    active = sources if sources else list(source_map.keys())
    session = make_session()
    api_results = []
    for name in active:
        if name in source_map:
            api_results.append(source_map[name](session, company))
    search_urls = generate_search_urls(company)
    entities = aggregate_entities(api_results)
    result = {
        "query": company,
        "search_date": datetime.now().isoformat(),
        "sources_searched": active,
        "total_from_apis": sum(r["count"] for r in api_results if isinstance(r.get("count"), int)),
        "entity_counts": {
            "parties": len(entities.get("parties", [])),
            "companies": len(entities.get("companies", [])),
            "law_firms": len(entities.get("law_firms", [])),
            "attorneys": len(entities.get("attorneys", [])),
            "judges": len(entities.get("judges", [])),
            "plaintiffs": len(entities.get("plaintiffs", [])),
        },
        "api_results": api_results,
        "entities": entities,
        "manual_search_urls": search_urls,
        "plaintiffs": entities.get("plaintiffs", []),
        "unique_plaintiff_count": len(entities.get("plaintiffs", [])),
    }
    _legal_cache[cache_key] = {"data": result, "ts": time.time()}
    return result


# ═════════════════════════════════════════════════════════════════════════
# Flask app
# ═════════════════════════════════════════════════════════════════════════

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024


# ── Home ─────────────────────────────────────────────────────────────────

@app.route("/")
def home():
    return render_template("home.html", active="")


# ── Legal Blueprint ──────────────────────────────────────────────────────

legal_bp = Blueprint("legal", __name__, url_prefix="/legal")


@legal_bp.route("/")
def legal_index():
    return render_template("legal.html", active="legal")


@legal_bp.route("/search")
def legal_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "No query"}), 400
    sources_param = request.args.get("sources", "").strip()
    sources = [s.strip() for s in sources_param.split(",") if s.strip()] if sources_param else None
    return jsonify(run_legal_search(q, sources=sources))


@legal_bp.route("/download/json")
def legal_download_json():
    q = request.args.get("q", "").strip()
    if not q:
        return "No query", 400
    data = run_legal_search(q)
    safe_name = re.sub(r'[^\w\-]', '_', q)
    path = os.path.join(tempfile.gettempdir(), f"legal_search_{safe_name}.json")
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    return send_file(path, as_attachment=True, download_name=f"legal_search_{q}.json",
                     mimetype="application/json")


@legal_bp.route("/download/xlsx")
def legal_download_xlsx():
    q = request.args.get("q", "").strip()
    if not q:
        return "No query", 400
    data = run_legal_search(q)
    safe_name = re.sub(r'[^\w\-]', '_', q)

    # Try using the existing json_to_excel.py for rich formatting
    json_path = os.path.join(tempfile.gettempdir(), f"legal_search_{safe_name}.json")
    xlsx_path = os.path.join(tempfile.gettempdir(), f"legal_search_{safe_name}.xlsx")
    with open(json_path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "json_to_excel.py")
    if os.path.exists(script):
        result = subprocess.run(
            [sys.executable, script, json_path, xlsx_path],
            capture_output=True, text=True,
        )
        if result.returncode == 0:
            return send_file(xlsx_path, as_attachment=True,
                             download_name=f"legal_search_{q}.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Fallback: build Excel inline
    sheets = []
    # Cases sheet
    case_rows = []
    for src in data.get("api_results", []):
        for c in src.get("cases", []):
            case_rows.append([
                src["source"],
                c.get("name", c.get("entity", "")),
                c.get("plaintiff", ""),
                c.get("defendant", ""),
                c.get("court", ""),
                c.get("date_filed", c.get("file_date", "")),
                c.get("docket_number", ""),
                c.get("file_description", ""),
            ])
    sheets.append({
        "title": "Cases",
        "headers": ["Source", "Case Name", "Plaintiff", "Defendant", "Court", "Date Filed", "Docket #", "Description"],
        "rows": case_rows,
    })
    # Parties sheet
    party_rows = [[p["name"], p.get("source", ""), p.get("case", "")]
                  for p in data.get("plaintiffs", [])]
    sheets.append({
        "title": "Plaintiffs & Parties",
        "headers": ["Name", "Source", "Case"],
        "rows": party_rows,
    })
    # Search URLs
    url_rows = [[s["source"], s["url"], s["note"]]
                for s in data.get("manual_search_urls", [])]
    sheets.append({
        "title": "Manual Search URLs",
        "headers": ["Source", "URL", "Notes"],
        "rows": url_rows,
    })

    path = _create_xlsx_from_data(sheets, f"legal_search_{safe_name}.xlsx")
    return send_file(path, as_attachment=True,
                     download_name=f"legal_search_{q}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


app.register_blueprint(legal_bp)


# ── Treasury Blueprint ───────────────────────────────────────────────────

treasury_bp = Blueprint("treasury", __name__, url_prefix="/treasury")


@treasury_bp.route("/")
def treasury_index():
    ok, msg = _check_api_key()
    return render_template("treasury.html", active="treasury", api_ok=ok, api_msg=msg)


@treasury_bp.route("/api/bt-companies")
def bt_companies_api():
    companies, btc_price, err = fetch_bt_companies()
    if err:
        return jsonify({"error": err}), 400
    formatted = format_btc_holdings(companies, btc_price)
    return jsonify({"companies": formatted, "btc_price": btc_price, "total": len(formatted)})


@treasury_bp.route("/api/people")
def treasury_people():
    ok, msg = _check_api_key()
    if not ok:
        return jsonify({"error": msg}), 400
    company = request.args.get("company", "").strip()
    domain = request.args.get("domain", "").strip()
    role = request.args.get("role", "all")
    if not company:
        return jsonify({"error": "Company name required"}), 400
    titles = TITLE_SETS.get(role, TREASURY_TITLES)
    people = find_people_at_companies(
        [{"name": company, "domain": domain}],
        titles,
        max_companies=1,
    )
    return jsonify({"people": people, "company": company, "titles_searched": len(titles)})


@treasury_bp.route("/api/download/json")
def treasury_download_json():
    ok, msg = _check_api_key()
    if not ok:
        return msg, 400
    company = request.args.get("company", "").strip()
    domain = request.args.get("domain", "").strip()
    symbol = request.args.get("symbol", "").strip()
    role = request.args.get("role", "all")
    if not company:
        return "Company name required", 400
    titles = TITLE_SETS.get(role, TREASURY_TITLES)
    people = find_people_at_companies([{"name": company, "domain": domain}], titles, max_companies=1)

    holdings = None
    bt_cos, btc_price, _ = fetch_bt_companies()
    if bt_cos:
        formatted = format_btc_holdings(bt_cos, btc_price)
        for co in formatted:
            if co["symbol"] == symbol or co["name"].lower() == company.lower():
                holdings = co
                break

    output = {
        "company": company, "symbol": symbol, "role_filter": role,
        "btc_price": btc_price, "holdings": holdings,
        "contacts": people, "total_contacts": len(people),
    }
    safe = re.sub(r'[^\w\-]', '_', company)
    path = os.path.join(tempfile.gettempdir(), f"treasury_{safe}.json")
    with open(path, "w") as f:
        json.dump(output, f, indent=2)
    return send_file(path, as_attachment=True, download_name=f"treasury_{safe}.json",
                     mimetype="application/json")


@treasury_bp.route("/api/download/xlsx")
def treasury_download_xlsx():
    ok, msg = _check_api_key()
    if not ok:
        return msg, 400
    company = request.args.get("company", "").strip()
    domain = request.args.get("domain", "").strip()
    symbol = request.args.get("symbol", "").strip()
    role = request.args.get("role", "all")
    if not company:
        return "Company name required", 400

    titles = TITLE_SETS.get(role, TREASURY_TITLES)
    people = find_people_at_companies([{"name": company, "domain": domain}], titles, max_companies=1)

    holdings = {}
    bt_cos, btc_price, _ = fetch_bt_companies()
    if bt_cos:
        formatted = format_btc_holdings(bt_cos, btc_price)
        for co in formatted:
            if co["symbol"] == symbol or co["name"].lower() == company.lower():
                holdings = co
                break

    rows = []
    for p in people:
        rows.append([
            p.get("name", ""),
            p.get("title", ""),
            p.get("company", ""),
            p.get("email", ""),
            p.get("phone", ""),
            p.get("linkedin_url", ""),
            f"{p.get('city', '')}, {p.get('state', '')}".strip(", "),
            p.get("country", ""),
            p.get("company_industry", ""),
            holdings.get("btc_holdings", ""),
            holdings.get("usd_value", ""),
            holdings.get("market_cap", ""),
            holdings.get("symbol", symbol),
        ])

    sheets = [{
        "title": f"Treasury Contacts - {company}"[:31],
        "headers": ["Name", "Title", "Company", "Email", "Phone", "LinkedIn",
                     "Location", "Country", "Industry",
                     "BTC Holdings", "USD Value", "Market Cap", "Ticker"],
        "rows": rows,
    }]

    safe = re.sub(r'[^\w\-]', '_', company)
    path = _create_xlsx_from_data(sheets, f"treasury_{safe}.xlsx")
    return send_file(path, as_attachment=True, download_name=f"treasury_{safe}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


app.register_blueprint(treasury_bp)


# ── Events Blueprint ─────────────────────────────────────────────────────

events_bp = Blueprint("events", __name__, url_prefix="/events")


@events_bp.route("/")
def events_index():
    ok, _ = _check_api_key()
    return render_template("events.html", active="events", api_ok=ok)


@events_bp.route("/scan")
def events_scan():
    url = request.args.get("url", "").strip()
    if not url:
        return jsonify({"error": "Event URL is required"}), 400

    enrich = request.args.get("enrich", "0") == "1"

    result, err = scrape_luma_event(url)
    if err:
        return jsonify({"error": err}), 400

    if not result:
        return jsonify({"error": "Could not extract event data from the provided URL"}), 400

    attendees = result["attendees"]
    enriched_count = 0

    if enrich and attendees:
        attendees, enriched_count = enrich_attendees_with_apollo(attendees)

    return jsonify({
        "event": result["event"],
        "attendees": attendees,
        "total_attendees": len(attendees),
        "enriched_count": enriched_count,
    })


@events_bp.route("/download/json")
def events_download_json():
    url = request.args.get("url", "").strip()
    if not url:
        return "Event URL required", 400

    result, err = scrape_luma_event(url)
    if err:
        return err, 400

    safe = re.sub(r'[^\w\-]', '_', result["event"].get("name", "event"))[:40]
    path = os.path.join(tempfile.gettempdir(), f"event_{safe}.json")
    with open(path, "w") as f:
        json.dump(result, f, indent=2, default=str)
    return send_file(path, as_attachment=True, download_name=f"event_{safe}.json",
                     mimetype="application/json")


@events_bp.route("/download/xlsx")
def events_download_xlsx():
    url = request.args.get("url", "").strip()
    enrich = request.args.get("enrich", "0") == "1"
    if not url:
        return "Event URL required", 400

    result, err = scrape_luma_event(url)
    if err:
        return err, 400

    attendees = result["attendees"]
    enriched_count = 0
    if enrich and attendees:
        ok, _ = _check_api_key()
        if ok:
            attendees, enriched_count = enrich_attendees_with_apollo(attendees)

    event = result["event"]
    rows = []
    for a in attendees:
        rows.append([
            event.get("name", ""),
            event.get("date", ""),
            event.get("location", ""),
            a.get("name", ""),
            a.get("title", a.get("role", "")),
            a.get("company", ""),
            a.get("email", ""),
            a.get("phone", ""),
            a.get("linkedin_url", ""),
            f"{a.get('city', '')}, {a.get('state', '')}".strip(", "),
            a.get("country", ""),
        ])

    sheets = [{
        "title": f"Event Attendees"[:31],
        "headers": ["Event Name", "Event Date", "Location",
                     "Attendee Name", "Title", "Company",
                     "Email", "Phone", "LinkedIn",
                     "Location", "Country"],
        "rows": rows,
    }]

    safe = re.sub(r'[^\w\-]', '_', event.get("name", "event"))[:40]
    path = _create_xlsx_from_data(sheets, f"event_{safe}.xlsx")
    return send_file(path, as_attachment=True, download_name=f"event_{safe}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


app.register_blueprint(events_bp)


# ── Health check (for Render) ─────────────────────────────────────────────

@app.route("/health")
def health_check():
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat()})


# ── Legacy route redirects ───────────────────────────────────────────────

@app.route("/apollo")
def apollo_redirect():
    return redirect("/treasury")


# ── Cache warming ────────────────────────────────────────────────────────

def _warm_cache():
    try:
        fetch_bt_companies()
    except Exception:
        pass

threading.Thread(target=_warm_cache, daemon=True).start()


# ── Main ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", sys.argv[1] if len(sys.argv) > 1 else 5055))
    print(f"\n  SDM Industry Search Engine")
    print(f"  http://localhost:{port}")
    print(f"  ─────────────────────────")
    print(f"  Sectors:")
    print(f"    /legal     — Legal & Litigation Intelligence")
    print(f"    /treasury  — Digital Asset Treasury & Finance")
    print(f"    /events    — Event Intelligence")
    ok, msg = _check_api_key()
    if not ok:
        print(f"\n  Note: Apollo API key not configured.")
        print(f"  Legal + Treasury list work without it; contact lookup needs .env setup.")
    print()
    app.run(host="0.0.0.0", port=port, debug=False)
